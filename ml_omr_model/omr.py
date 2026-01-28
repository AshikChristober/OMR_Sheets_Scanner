from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import cv2
import numpy as np
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)
CORS(app)  # Allow React frontend access (localhost:3000)

# -------- Folder Setup --------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
REPORT_FOLDER = os.path.join(BASE_DIR, "reports")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

TASK_FILE = os.path.join(UPLOAD_FOLDER, "tasks.json")

# -------- Helper Functions --------
def load_tasks():
    if not os.path.exists(TASK_FILE):
        return []
    with open(TASK_FILE, "r") as f:
        return json.load(f)

def save_tasks(tasks):
    with open(TASK_FILE, "w") as f:
        json.dump(tasks, f, indent=2)

# -------- Routes --------
@app.route('/')
def home():
    return jsonify({"message": "OMR Backend is Running 🚀"})

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/reports/<filename>')
def get_report(filename):
    return send_from_directory(REPORT_FOLDER, filename)

@app.route('/save_answer_key', methods=['POST'])
def save_answer_key():
    data = request.get_json()
    group = data.get("group")
    answer_key = data.get("answer_key")

    if not group or not answer_key:
        return jsonify({"error": "Missing data"}), 400

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"answer_key_{group.replace(' ', '_').lower()}_{timestamp}.json"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    with open(filepath, "w") as f:
        json.dump(answer_key, f, indent=2)

    # Save task details
    tasks = load_tasks()
    new_task = {
        "id": len(tasks) + 1,
        "category": group,
        "file_path": filepath,
        "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    tasks.append(new_task)
    save_tasks(tasks)

    return jsonify({
        "message": f"Task created for {group} ✅",
        "task": new_task
    })

@app.route('/get_tasks', methods=['GET'])
def get_tasks():
    return jsonify(load_tasks())

# -------- Image Processing --------
def process_image(file_path, custom_key=None):
    try:
        img = cv2.imread(file_path)
        if img is None:
            return {"error": "Image not found or unreadable"}

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        bubbles = [c for c in contours if 400 < cv2.contourArea(c) < 2000]
        bubbles = sorted(bubbles, key=lambda c: cv2.boundingRect(c)[1])

        answer_key = custom_key if custom_key else {}
        results, correct_count, question_number = {}, 0, 1
        options = ['A', 'B', 'C', 'D']

        for i in range(0, len(bubbles), 4):
            group = sorted(bubbles[i:i + 4], key=lambda c: cv2.boundingRect(c)[0])
            fill_intensity = []
            for b in group:
                mask = np.zeros(thresh.shape, dtype="uint8")
                cv2.drawContours(mask, [b], -1, 255, -1)
                mean_val = cv2.mean(thresh, mask=mask)[0]
                fill_intensity.append(mean_val)

            filled_index = np.argmax(fill_intensity)
            marked_option = options[filled_index]
            correct_option = answer_key.get(str(question_number))
            is_correct = marked_option == correct_option

            if is_correct:
                correct_count += 1

            results[question_number] = {
                "marked": marked_option,
                "correct": correct_option,
                "is_correct": is_correct
            }
            question_number += 1

        total_q = len(answer_key) if len(answer_key) > 0 else 1
        score = round((correct_count / total_q) * 100, 2)

        debug_image = img.copy()
        cv2.drawContours(debug_image, bubbles, -1, (0, 255, 0), 2)
        debug_path = os.path.join(UPLOAD_FOLDER, f"debug_{os.path.basename(file_path)}")
        cv2.imwrite(debug_path, debug_image)

        return {
            "status": "processed",
            "total_questions": total_q,
            "correct_answers": correct_count,
            "score_percent": score,
            "details": results,
            "debug_image": f"/uploads/{os.path.basename(debug_path)}"
        }

    except Exception as e:
        return {"error": str(e)}

# -------- Evaluation & Report Generation --------
latest_results = {}

@app.route('/evaluate_task/<int:task_id>', methods=['POST'])
def evaluate_task(task_id):
    tasks = load_tasks()
    task = next((t for t in tasks if t["id"] == task_id), None)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    filename = f"task_{task_id}_{file.filename}"
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)

    with open(task["file_path"]) as f:
        answer_key = json.load(f)

    result = process_image(file_path, custom_key=answer_key)
    latest_results[task_id] = result

    return jsonify({
        "message": f"Evaluation completed for Task {task_id}",
        "task": task,
        "result": result
    })

@app.route('/generate_report/<int:task_id>', methods=['GET'])
def generate_report(task_id):
    if task_id not in latest_results:
        return jsonify({"error": "No recent evaluation found for this task"}), 400

    result = latest_results[task_id]
    details = result["details"]

    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Evaluation Report"

    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    # Headers
    headers = ["Question No.", "Marked Option", "Correct Option", "Status"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data rows
    for q_no, data in details.items():
        status = "Correct" if data["is_correct"] else "Wrong"
        row = [q_no, data["marked"], data["correct"], status]
        ws.append(row)
        fill = correct_fill if data["is_correct"] else wrong_fill
        for col_num in range(1, 5):
            ws.cell(row=ws.max_row, column=col_num).fill = fill
            ws.cell(row=ws.max_row, column=col_num).alignment = center_align

    # Summary section
    ws.append([])
    ws.append(["", "", "Total Questions", result["total_questions"]])
    ws.append(["", "", "Correct Answers", result["correct_answers"]])
    ws.append(["", "", "Score (%)", result["score_percent"]])

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"report_task_{task_id}_{timestamp}.xlsx"
    report_path = os.path.join(REPORT_FOLDER, report_filename)
    wb.save(report_path)

    return jsonify({
        "message": "Report generated successfully!",
        "file_url": f"/reports/{report_filename}"
    })

if __name__ == '__main__':
    app.run(debug=True)
